import pyautogui as pg
from time import sleep
import subprocess
import openpyxl

# Variables
nav = 'Safari'
dados = openpyxl.load_workbook('dados.xlsx')
folha = dados['Folha 1']

# receivers (I've made with 5 contacts, 
# but it can be for as many as you need)
d1 = folha.cell(row=2, column=1).value
d2 = folha.cell(row=3, column=1).value
d3 = folha.cell(row=4, column=1).value
d4 = folha.cell(row=5, column=1).value
d5 = folha.cell(row=6, column=1).value

# messages (I've made that you can write
# one exclusive message for each contact,
# but you also can use the same message
# for every receiver)
m1 = folha.cell(row=2, column=2).value
m2 = folha.cell(row=3, column=2).value
m3 = folha.cell(row=4, column=2).value
m4 = folha.cell(row=5, column=2).value
m5 = folha.cell(row=6, column=2).value

# Functions 
# 1 - TAB (for finding the searchbox)
def Tab1():
    try:
        for c in range(1, 7):
            pg.press('tab')
        print('searchbox founded')
    except Exception as e:
        print(f'error: {e}')
def Tab2():
    try:
        for c in range(1, 8):
            pg.press('tab')
        print('searchbox founded again')
    except Exception as e:
        print(f'error: {e}')

# 2 - CONTACTS (for writing the contacts)

def D1():
    try:
        pg.write(d1)
        pg.press('enter')
        print(f'{d1} founded successfully')
    except Exception as e:
        print(f'error: {e}')
def D2():
    try:
        pg.write(d2)
        pg.press('enter')
        print(f'{d2} founded successfully')
    except Exception as e:
        print(f'error: {e}')
def D3():
    try:
        pg.write(d3)
        pg.press('enter')
        print(f'{d3} founded successfully')
    except Exception as e:
        print(f'error: {e}')
def D4():
    try:
        pg.write(d4)
        pg.press('enter')
        print(f'{d4} founded successfully')
    except Exception as e:
        print(f'error: {e}')
def D5():
    try:
        pg.write(d5)
        pg.press('enter')
        print(f'{d5} founded successfully')
    except Exception as e:
        print(f'error: {e}')

# 3 - MESSAGES (for writing the messages for each contact)
# still don't know how to encode texts in python so the
# accentuation may be a little bit funky

def M1():
    try:
        pg.write(m1)
        pg.press('enter')
        print('message sended successfuly')
    except Exception as e:
        print(f'error: {e}')
def M2():
    try:
        pg.write(m2)
        pg.press('enter')
        print('message sended successfuly')
    except Exception as e:
        print(f'error: {e}')
def M3():
    try:
        pg.write(m3)
        pg.press('enter')
        print('message sended successfuly')
    except Exception as e:
        print(f'error: {e}')
def M4():
    try:
        pg.write(m4)
        pg.press('enter')
        print('message sended successfuly')
    except Exception as e:
        print(f'error: {e}')
def M5():
    try:
        pg.write(m5)
        pg.press('enter')
        print('message sended successfuly')
    except Exception as e:
        print(f'error: {e}')


# Open WhatsApp in a new tab in safari
try:
    subprocess.Popen(['open', '-a', nav])
    sleep(5)
    pg.keyDown('command')
    pg.press('t')
    pg.keyUp('command')
    sleep(2)
    pg.write('https://web.whatsapp.com')
    pg.press('enter')
    print(f'New tab in {nav} opened successfuly')
except Exception as e:
    print(f'error {e}')

sleep(20)
Tab1()
sleep(2)
D1()
sleep(2)
M1()

# the scritp will follow the same logic as long as you need
# if you need to send more messages for more people, just
# keep going e adjust the amount of commands and functions